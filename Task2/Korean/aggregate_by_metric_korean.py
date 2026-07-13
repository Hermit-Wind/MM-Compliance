#!/usr/bin/env python3
"""
Aggregate evaluation results by metric across companies - Korean version

This script reads Korean result files and creates Excel tables
where each row represents a metric-company combination.

Outputs TWO sets of files based on complete mode:
- Mixed mode (complete=TRUE + FALSE):
  - results_by_metric_full_mixed.xlsx (with nDCG metrics)
  - results_by_metric_no_ndcg_mixed.xlsx (without nDCG metrics)
- Complete=YES mode:
  - results_by_metric_full_complete_yes.xlsx (with nDCG metrics)
  - results_by_metric_no_ndcg_complete_yes.xlsx (without nDCG metrics)
"""

import json
from pathlib import Path
from collections import defaultdict
import pandas as pd
import argparse


def get_company_name(cid: str) -> str:
    """
    Get company name from CID
    For Korean data, CID is the PDF stem
    
    Args:
        cid: PDF filename like "bh.pdf"
    
    Returns:
        Company name like "bh"
    """
    return Path(cid).stem


# =====================================================
# Main Processing Function
# =====================================================
def aggregate_results_by_metric(results_dir: Path, output_dir: Path):
    """
    Aggregate results by metric across all companies
    Processes both mixed and complete_yes modes separately
    
    Args:
        results_dir: Directory containing result files (organized by industry/mode)
        output_dir: Directory to save output Excel files
    """
    print(f"\n{'='*80}")
    print(f"AGGREGATING KOREAN RESULTS BY METRIC")
    print(f"{'='*80}")
    print(f"Input directory: {results_dir}")
    print(f"Output directory: {output_dir}")
    
    # Collect data separately by mode
    data_by_mode = {
        'mixed': [],
        'complete_yes': []
    }
    
    # Find all result JSON files in subdirectories (industry/mode structure)
    result_files = []
    for industry_dir in results_dir.iterdir():
        if industry_dir.is_dir():
            for mode_dir in industry_dir.iterdir():
                if mode_dir.is_dir():
                    result_files.extend(list(mode_dir.glob("results_*.json")))
    
    if not result_files:
        print(f"\n❌ Error: No result files found in {results_dir}")
        print(f"   Looking for files matching pattern: */*/results_*.json")
        return
    
    print(f"\nFound {len(result_files)} result files:")
    for rf in sorted(result_files):
        print(f"   • {rf.parent.parent.name}/{rf.parent.name}/{rf.name}")
    
    # Process each result file
    for result_file in sorted(result_files):
        # Determine mode from parent directory or filename
        if "mixed" in result_file.parent.name or "mixed" in result_file.name:
            mode = "mixed"
        elif "complete_yes" in result_file.parent.name or "complete_yes" in result_file.name:
            mode = "complete_yes"
        else:
            print(f"⚠️  Skipping {result_file.name}: Cannot determine mode")
            continue
        
        # Get industry from parent's parent
        industry_name = result_file.parent.parent.name
        
        print(f"\n📂 Processing {industry_name}/{result_file.parent.name}/{result_file.name} (mode: {mode}):")
        
        # Load results
        with open(result_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        model = data.get('model', 'Unknown')
        language = data.get('language', 'korean')
        total_queries = len(data.get('results', []))
        
        print(f"   Model: {model}")
        print(f"   Language: {language}")
        print(f"   Total queries: {total_queries}")
        
        # Process each query result
        for result in data.get('results', []):
            cid = result.get('CID', '')
            
            # Get company name from CID
            company_name = get_company_name(cid)
            
            # Extract all relevant fields
            row = {
                'Metric': result.get('metric', ''),
                'Code': result.get('code', ''),
                'Topic': result.get('topic', ''),
                'Company': company_name,
                'Industry': result.get('industry', industry_name),
                'Language': language,
                'CID': cid,
                'Model': model,
                'Complete_Flag': result.get('complete_flag', ''),
                'Num_Relevant_Pages': result.get('num_relevant_pages', 0),
            }
            
            # Add all evaluation metrics
            metrics = result.get('metrics', {})
            
            # Add Recall metrics
            row['Recall@1'] = metrics.get('Recall@1', None)
            row['Recall@5'] = metrics.get('Recall@5', None)
            row['Recall@10'] = metrics.get('Recall@10', None)
            row['Recall@20'] = metrics.get('Recall@20', None)
            
            # Add nDCG metrics (will be removed in no_ndcg version)
            row['nDCG@1'] = metrics.get('nDCG@1', None)
            row['nDCG@5'] = metrics.get('nDCG@5', None)
            row['nDCG@10'] = metrics.get('nDCG@10', None)
            row['nDCG@20'] = metrics.get('nDCG@20', None)
            
            # Add Precision metrics
            row['Precision@1'] = metrics.get('Precision@1', None)
            row['Precision@5'] = metrics.get('Precision@5', None)
            row['Precision@10'] = metrics.get('Precision@10', None)
            row['Precision@20'] = metrics.get('Precision@20', None)
            
            # Add Hit metrics
            row['Hit@1'] = metrics.get('Hit@1', None)
            row['Hit@5'] = metrics.get('Hit@5', None)
            row['Hit@10'] = metrics.get('Hit@10', None)
            row['Hit@20'] = metrics.get('Hit@20', None)
            
            # Add to appropriate mode list
            data_by_mode[mode].append(row)
    
    # Process each mode separately
    output_dir.mkdir(parents=True, exist_ok=True)
    
    for mode, mode_data in data_by_mode.items():
        if not mode_data:
            print(f"\n⚠️  No data found for mode: {mode}")
            continue
        
        print(f"\n{'='*80}")
        print(f"PROCESSING MODE: {mode.upper()}")
        print(f"{'='*80}")
        print(f"✓ Collected {len(mode_data)} metric-company combinations")
        
        # Create DataFrame
        df_full = pd.DataFrame(mode_data)
        
        # Sort by Metric, then Company for easy comparison
        df_full = df_full.sort_values(['Metric', 'Company']).reset_index(drop=True)
        
        # Define column order for full version
        columns_full = [
            'Metric', 'Code', 'Topic', 'Company', 'Industry', 'Language', 'CID', 'Model', 
            'Complete_Flag', 'Num_Relevant_Pages',
            'Recall@1', 'Recall@5', 'Recall@10', 'Recall@20',
            'nDCG@1', 'nDCG@5', 'nDCG@10', 'nDCG@20',
            'Precision@1', 'Precision@5', 'Precision@10', 'Precision@20',
            'Hit@1', 'Hit@5', 'Hit@10', 'Hit@20'
        ]
        
        # Reorder columns
        df_full = df_full[columns_full]
        
        # Create no-nDCG version by removing nDCG columns
        columns_no_ndcg = [col for col in columns_full if not col.startswith('nDCG')]
        df_no_ndcg = df_full[columns_no_ndcg].copy()
        
        # Print statistics
        print(f"\n📊 Statistics for {mode}:")
        print(f"   Total rows: {len(df_full)}")
        print(f"   Unique metrics: {df_full['Metric'].nunique()}")
        print(f"   Unique companies: {df_full['Company'].nunique()}")
        print(f"   Industries: {df_full['Industry'].nunique()}")
        
        # Show company list
        print(f"\n   Companies found:")
        for company in sorted(df_full['Company'].unique()):
            count = len(df_full[df_full['Company'] == company])
            print(f"      • {company}: {count} queries")
        
        # Show metrics distribution
        print(f"\n   Metrics by company count:")
        metric_counts = df_full.groupby('Metric')['Company'].nunique().value_counts().sort_index()
        for count, num_metrics in metric_counts.items():
            print(f"      • {num_metrics} metrics appear in {count} company(ies)")
        
        # Define output filenames
        output_file_full = output_dir / f"results_by_metric_full_{mode}.xlsx"
        output_file_no_ndcg = output_dir / f"results_by_metric_no_ndcg_{mode}.xlsx"
        
        print(f"\n💾 Saving Excel files for {mode}...")
        
        # Save full version
        with pd.ExcelWriter(output_file_full, engine='openpyxl') as writer:
            df_full.to_excel(writer, sheet_name='Results', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Results']
            for idx, col in enumerate(df_full.columns):
                max_length = max(
                    df_full[col].astype(str).map(len).max(),
                    len(col)
                )
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        print(f"   ✓ {output_file_full.name}")
        print(f"      Columns: {len(df_full.columns)}")
        print(f"      Rows: {len(df_full)}")
        
        # Save no-nDCG version
        with pd.ExcelWriter(output_file_no_ndcg, engine='openpyxl') as writer:
            df_no_ndcg.to_excel(writer, sheet_name='Results', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Results']
            for idx, col in enumerate(df_no_ndcg.columns):
                max_length = max(
                    df_no_ndcg[col].astype(str).map(len).max(),
                    len(col)
                )
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        print(f"   ✓ {output_file_no_ndcg.name}")
        print(f"      Columns: {len(df_no_ndcg.columns)}")
        print(f"      Rows: {len(df_no_ndcg)}")
        
        # Show sample data
        print(f"\n📋 Sample data for {mode} (first 5 rows):")
        print(df_full[['Metric', 'Company', 'Industry', 'Recall@1', 'Recall@5']].head().to_string())
    
    # Final summary
    print(f"\n{'='*80}")
    print(f"✅ DONE!")
    print(f"{'='*80}")
    print(f"\n📁 Output files created:")
    
    for mode in ['mixed', 'complete_yes']:
        if data_by_mode[mode]:
            print(f"\n   {mode.upper()}:")
            print(f"      • results_by_metric_full_{mode}.xlsx")
            print(f"      • results_by_metric_no_ndcg_{mode}.xlsx")
    
    print(f"\n   Output directory: {output_dir}")
    print(f"{'='*80}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Aggregate Korean evaluation results by metric across companies (split by complete mode)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process default Korean results
  python aggregate_by_metric_korean.py
  
  # Process with custom input
  python aggregate_by_metric_korean.py --input results_colqwen_eval_korean
  
  # Process with custom output
  python aggregate_by_metric_korean.py --output analysis_korean

Output (TWO sets based on complete mode):
  Mixed mode (complete=TRUE + FALSE):
    - results_by_metric_full_mixed.xlsx (with nDCG@k metrics)
    - results_by_metric_no_ndcg_mixed.xlsx (without nDCG@k metrics)
  
  Complete=YES mode:
    - results_by_metric_full_complete_yes.xlsx (with nDCG@k metrics)
    - results_by_metric_no_ndcg_complete_yes.xlsx (without nDCG@k metrics)
        """
    )
    
    parser.add_argument(
        '--input',
        default='results_colqwen_eval_korean',
        help='Input results directory (default: results_colqwen_eval_korean)'
    )
    
    parser.add_argument(
        '--output',
        default='metric_analysis_korean',
        help='Output directory for Excel files (default: metric_analysis_korean)'
    )
    
    args = parser.parse_args()
    
    # Setup paths
    base_dir = Path(__file__).resolve().parent
    results_dir = base_dir / args.input
    output_dir = base_dir / args.output
    
    # Check if input directory exists
    if not results_dir.exists():
        print(f"\n❌ Error: Input directory not found: {results_dir}")
        print(f"   Please check the directory name and try again.")
        return
    
    # Run aggregation
    aggregate_results_by_metric(results_dir, output_dir)


if __name__ == "__main__":
    main()