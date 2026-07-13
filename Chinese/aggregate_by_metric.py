#!/usr/bin/env python3
"""
Aggregate evaluation results by metric across companies - Chinese version with industry support

This script reads Chinese result files (organized by industry) and creates Excel tables
where each row represents a metric-company combination.

Outputs:
- results_by_metric_full.xlsx (with nDCG metrics)
- results_by_metric_no_ndcg.xlsx (without nDCG metrics)
"""

import json
from pathlib import Path
from collections import defaultdict
import pandas as pd
import argparse


def get_company_name(cid: str) -> str:
    """
    Get company name from CID
    For Chinese data, CID is already the company name (PDF stem)
    
    Args:
        cid: PDF filename like "Alchip.pdf"
    
    Returns:
        Company name like "Alchip"
    """
    return Path(cid).stem


def aggregate_results_by_metric(input_dir: Path, output_dir: Path, industries: list = None):
    """
    Aggregate results by metric across all companies and industries
    
    Args:
        input_dir: Base directory containing industry subdirectories
        output_dir: Directory to save output Excel files
        industries: List of specific industries to process (None = all)
    """
    print(f"\n{'='*80}")
    print(f"AGGREGATING CHINESE RESULTS BY METRIC")
    print(f"{'='*80}")
    print(f"Input directory: {input_dir}")
    print(f"Output directory: {output_dir}")
    
    if not input_dir.exists():
        print(f"\n❌ Error: Input directory not found: {input_dir}")
        return
    
    # Find all industry subdirectories
    industry_dirs = [d for d in input_dir.iterdir() if d.is_dir()]
    
    if not industry_dirs:
        print(f"\n❌ Error: No industry directories found in {input_dir}")
        return
    
    # Filter industries if specified
    if industries:
        industry_dirs = [d for d in industry_dirs if d.name in industries]
        if not industry_dirs:
            print(f"\n❌ Error: None of the specified industries found")
            return
    
    print(f"\n📂 Found {len(industry_dirs)} industry directories to process:")
    for d in sorted(industry_dirs):
        print(f"   • {d.name}")
    
    # Collect all data across industries
    all_data = []
    
    for industry_dir in sorted(industry_dirs):
        industry = industry_dir.name
        result_file = industry_dir / f"results_{industry}.json"
        
        if not result_file.exists():
            print(f"\n⚠ Warning: Result file not found for {industry}: {result_file}")
            continue
        
        print(f"\n📄 Processing {industry}...")
        
        # Load results
        with open(result_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        model = data.get('model', 'Unknown')
        language = data.get('language', 'chinese')
        total_queries = len(data.get('results', []))
        
        print(f"   Model: {model}")
        print(f"   Language: {language}")
        print(f"   Queries: {total_queries}")
        
        # Process each query result
        for result in data.get('results', []):
            cid = result.get('CID', '')
            
            # Get company name from CID
            company_name = get_company_name(cid)
            
            # Extract all relevant fields
            row = {
                'Metric': result.get('metric', ''),
                'Code': result.get('code', ''),
                'Topic': result.get('topic', ''),
                'Company': company_name,
                'Industry': result.get('industry', industry),
                'Language': language,
                'CID': cid,
                'Model': model,
                'Num_Relevant_Pages': result.get('num_relevant_pages', 0),
            }
            
            # Add all evaluation metrics
            metrics = result.get('metrics', {})
            
            # Add Recall metrics
            row['Recall@1'] = metrics.get('Recall@1', None)
            row['Recall@5'] = metrics.get('Recall@5', None)
            row['Recall@10'] = metrics.get('Recall@10', None)
            row['Recall@20'] = metrics.get('Recall@20', None)
            
            # Add nDCG metrics (will be removed in no_ndcg version)
            row['nDCG@1'] = metrics.get('nDCG@1', None)
            row['nDCG@5'] = metrics.get('nDCG@5', None)
            row['nDCG@10'] = metrics.get('nDCG@10', None)
            row['nDCG@20'] = metrics.get('nDCG@20', None)
            
            # Add Precision metrics
            row['Precision@1'] = metrics.get('Precision@1', None)
            row['Precision@5'] = metrics.get('Precision@5', None)
            row['Precision@10'] = metrics.get('Precision@10', None)
            row['Precision@20'] = metrics.get('Precision@20', None)
            
            # Add Hit metrics
            row['Hit@1'] = metrics.get('Hit@1', None)
            row['Hit@5'] = metrics.get('Hit@5', None)
            row['Hit@10'] = metrics.get('Hit@10', None)
            row['Hit@20'] = metrics.get('Hit@20', None)
            
            all_data.append(row)
        
        print(f"   ✓ Collected {total_queries} metric-company combinations")
    
    print(f"\n✓ Total collected: {len(all_data)} metric-company combinations")
    
    if len(all_data) == 0:
        print(f"\n❌ Error: No data collected. Please check your input files.")
        return
    
    # Create DataFrame
    df_full = pd.DataFrame(all_data)
    
    # Sort by Industry, then Metric, then Company for easy comparison
    df_full = df_full.sort_values(['Industry', 'Metric', 'Company']).reset_index(drop=True)
    
    # Define column order for full version (with Industry after Company)
    columns_full = [
        'Metric', 'Code', 'Topic', 'Company', 'Industry', 'Language', 'CID', 'Model', 'Num_Relevant_Pages',
        'Recall@1', 'Recall@5', 'Recall@10', 'Recall@20',
        'nDCG@1', 'nDCG@5', 'nDCG@10', 'nDCG@20',
        'Precision@1', 'Precision@5', 'Precision@10', 'Precision@20',
        'Hit@1', 'Hit@5', 'Hit@10', 'Hit@20'
    ]
    
    # Reorder columns
    df_full = df_full[columns_full]
    
    # Create no-nDCG version by removing nDCG columns
    columns_no_ndcg = [col for col in columns_full if not col.startswith('nDCG')]
    df_no_ndcg = df_full[columns_no_ndcg].copy()
    
    # Print statistics
    print(f"\n📊 Statistics:")
    print(f"   Total rows: {len(df_full)}")
    print(f"   Unique metrics: {df_full['Metric'].nunique()}")
    print(f"   Unique companies: {df_full['Company'].nunique()}")
    print(f"   Unique industries: {df_full['Industry'].nunique()}")
    
    # Show industry breakdown
    print(f"\n   Breakdown by industry:")
    for industry in sorted(df_full['Industry'].unique()):
        count = len(df_full[df_full['Industry'] == industry])
        companies = df_full[df_full['Industry'] == industry]['Company'].nunique()
        print(f"      • {industry}: {count} queries, {companies} companies")
    
    # Show company list
    print(f"\n   Companies by industry:")
    for industry in sorted(df_full['Industry'].unique()):
        companies = sorted(df_full[df_full['Industry'] == industry]['Company'].unique())
        print(f"      • {industry}: {', '.join(companies)}")
    
    # Show metrics distribution
    print(f"\n   Metrics by company count:")
    metric_counts = df_full.groupby('Metric')['Company'].nunique().value_counts().sort_index()
    for count, num_metrics in metric_counts.items():
        print(f"      • {num_metrics} metrics appear in {count} company(ies)")
    
    # Save to Excel files
    output_dir.mkdir(parents=True, exist_ok=True)
    
    output_file_full = output_dir / "results_by_metric_full.xlsx"
    output_file_no_ndcg = output_dir / "results_by_metric_no_ndcg.xlsx"
    
    print(f"\n💾 Saving Excel files...")
    
    # Save full version
    with pd.ExcelWriter(output_file_full, engine='openpyxl') as writer:
        df_full.to_excel(writer, sheet_name='Results', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Results']
        for idx, col in enumerate(df_full.columns):
            max_length = max(
                df_full[col].astype(str).map(len).max(),
                len(col)
            )
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    print(f"   ✓ {output_file_full.name}")
    print(f"      Columns: {len(df_full.columns)}")
    print(f"      Rows: {len(df_full)}")
    
    # Save no-nDCG version
    with pd.ExcelWriter(output_file_no_ndcg, engine='openpyxl') as writer:
        df_no_ndcg.to_excel(writer, sheet_name='Results', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Results']
        for idx, col in enumerate(df_no_ndcg.columns):
            max_length = max(
                df_no_ndcg[col].astype(str).map(len).max(),
                len(col)
            )
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    print(f"   ✓ {output_file_no_ndcg.name}")
    print(f"      Columns: {len(df_no_ndcg.columns)}")
    print(f"      Rows: {len(df_no_ndcg)}")
    
    # Show sample data
    print(f"\n📋 Sample data (first 5 rows):")
    print(df_full[['Metric', 'Company', 'Industry', 'Recall@1', 'Recall@5']].head().to_string())
    
    print(f"\n✅ DONE!")
    print(f"   Full version: {output_file_full}")
    print(f"   No-nDCG version: {output_file_no_ndcg}")
    print(f"{'='*80}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Aggregate Chinese evaluation results by metric across companies and industries",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process all industries
  python aggregate_by_metric_chinese.py
  
  # Process specific industries only
  python aggregate_by_metric_chinese.py --industries semiconductor finance
  
  # Process with custom input directory
  python aggregate_by_metric_chinese.py --input results_colqwen_eval_chinese
  
  # Process with custom output directory
  python aggregate_by_metric_chinese.py --output analysis_chinese

Output:
  - results_by_metric_full.xlsx (with nDCG@k metrics)
  - results_by_metric_no_ndcg.xlsx (without nDCG@k metrics)
  
Directory structure:
  Input:  results_colqwen_eval_chinese/
          ├── semiconductor/results_semiconductor.json
          ├── finance/results_finance.json
          └── energy/results_energy.json
        """
    )
    
    parser.add_argument(
        '--input',
        default='results_colqwen_eval_chinese',
        help='Input base directory containing industry subdirectories (default: results_colqwen_eval_chinese)'
    )
    
    parser.add_argument(
        '--output',
        default='metric_analysis_chinese',
        help='Output directory for Excel files (default: metric_analysis_chinese)'
    )
    
    parser.add_argument(
        '--industries',
        nargs='+',
        default=None,
        help='Specific industries to process (default: all available)'
    )
    
    args = parser.parse_args()
    
    # Setup paths
    base_dir = Path(__file__).resolve().parent
    input_dir = base_dir / args.input
    output_dir = base_dir / args.output
    
    # Run aggregation
    aggregate_results_by_metric(input_dir, output_dir, args.industries)


if __name__ == "__main__":
    main()