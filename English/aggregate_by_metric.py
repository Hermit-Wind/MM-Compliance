#!/usr/bin/env python3
"""
Aggregate evaluation results by metric across companies

This script reads original result files (before company split) and creates Excel tables
where each row represents a metric-company combination.

Outputs:
- results_by_metric_full.xlsx (with nDCG metrics)
- results_by_metric_no_ndcg.xlsx (without nDCG metrics)
"""

import json
from pathlib import Path
from collections import defaultdict
import pandas as pd
import argparse


# =====================================================
# CID to Company Name Mapping
# =====================================================
CID_TO_COMPANY = {
    # Exact matches (with or without .pdf extension)
    "06-21_gb-_lvmh_rse2022.pdf": "LVMH",
    "06-21_gb-_lvmh_rse2022": "LVMH",
    
    "2022-Baytex-ESG-Report-FINAL.pdf": "Baytex Energy",
    "2022-Baytex-ESG-Report-FINAL": "Baytex Energy",
    
    "Burberry_2020-21_ESG.pdf": "Burberry",
    "Burberry_2020-21_ESG": "Burberry",
    
    "CanNickel-ESG-Report2023-print.pdf": "Canada Nickel",
    "CanNickel-ESG-Report2023-print": "Canada Nickel",
    
    "esg-2022.pdf": "Greenergy",
    "esg-2022": "Greenergy",
    
    "esg-report-2022.pdf": "Arab Bank",
    "esg-report-2022": "Arab Bank",
    
    "ESG-Report-2022-Final_ADA.pdf": "U.S. Bancorp",
    "ESG-Report-2022-Final_ADA": "U.S. Bancorp",
    
    "ESGReport2022.pdf": "Standard Bank",
    "ESGReport2022": "Standard Bank",
    
    "Kering_Sustainability_Progress_Report_2020_2023_7d06687606.pdf": "Kering",
    "Kering_Sustainability_Progress_Report_2020_2023_7d06687606": "Kering",
}


def get_company_name(cid: str) -> str:
    """
    Get company name from CID using mapping table
    
    Args:
        cid: PDF filename like "2022-Baytex-ESG-Report-FINAL.pdf"
    
    Returns:
        Company name like "Baytex Energy"
    """
    # Try exact match first (with .pdf)
    if cid in CID_TO_COMPANY:
        return CID_TO_COMPANY[cid]
    
    # Try without .pdf extension
    cid_stem = Path(cid).stem
    if cid_stem in CID_TO_COMPANY:
        return CID_TO_COMPANY[cid_stem]
    
    # Fallback: return CID itself
    print(f"⚠️  Warning: No mapping found for CID '{cid}'")
    return cid


# =====================================================
# Main Processing Function
# =====================================================
def aggregate_results_by_metric(results_dir: Path, output_dir: Path):
    """
    Aggregate results by metric across all companies
    
    Args:
        results_dir: Directory containing original result files (results_energy.json, etc.)
        output_dir: Directory to save output Excel files
    """
    print(f"\n{'='*80}")
    print(f"AGGREGATING RESULTS BY METRIC")
    print(f"{'='*80}")
    print(f"Input directory: {results_dir}")
    print(f"Output directory: {output_dir}")
    
    # Collect all data
    all_data = []
    
    # Find all result JSON files
    result_files = list(results_dir.glob("results_*.json"))
    
    if not result_files:
        print(f"\n❌ Error: No result files found in {results_dir}")
        print(f"   Looking for files matching pattern: results_*.json")
        return
    
    print(f"\nFound {len(result_files)} result files:")
    for rf in sorted(result_files):
        print(f"   • {rf.name}")
    
    # Process each result file
    for result_file in sorted(result_files):
        industry_name = result_file.stem.replace('results_', '')
        
        print(f"\n📂 Processing {result_file.name} (industry: {industry_name}):")
        
        # Load results
        with open(result_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        model = data.get('model', 'Unknown')
        total_queries = len(data.get('results', []))
        
        print(f"   Model: {model}")
        print(f"   Total queries: {total_queries}")
        
        # Process each query result
        for result in data.get('results', []):
            cid = result.get('CID', '')
            
            # Get company name from CID
            company_name = get_company_name(cid)
            
            # Extract all relevant fields
            row = {
                'Metric': result.get('metric', ''),
                'Code': result.get('code', ''),
                'Topic': result.get('topic', ''),
                'Company': company_name,
                'Industry': industry_name,
                'CID': cid,
                'Model': model,
                'Num_Relevant_Pages': result.get('num_relevant_pages', 0),
            }
            
            # Add all evaluation metrics
            metrics = result.get('metrics', {})
            
            # Add Recall metrics
            row['Recall@1'] = metrics.get('Recall@1', None)
            row['Recall@5'] = metrics.get('Recall@5', None)
            row['Recall@10'] = metrics.get('Recall@10', None)
            row['Recall@20'] = metrics.get('Recall@20', None)
            
            # Add nDCG metrics (will be removed in no_ndcg version)
            row['nDCG@1'] = metrics.get('nDCG@1', None)
            row['nDCG@5'] = metrics.get('nDCG@5', None)
            row['nDCG@10'] = metrics.get('nDCG@10', None)
            row['nDCG@20'] = metrics.get('nDCG@20', None)
            
            # Add Precision metrics
            row['Precision@1'] = metrics.get('Precision@1', None)
            row['Precision@5'] = metrics.get('Precision@5', None)
            row['Precision@10'] = metrics.get('Precision@10', None)
            row['Precision@20'] = metrics.get('Precision@20', None)
            
            # Add Hit metrics
            row['Hit@1'] = metrics.get('Hit@1', None)
            row['Hit@5'] = metrics.get('Hit@5', None)
            row['Hit@10'] = metrics.get('Hit@10', None)
            row['Hit@20'] = metrics.get('Hit@20', None)
            
            all_data.append(row)
    
    print(f"\n✓ Collected {len(all_data)} metric-company combinations")
    
    if len(all_data) == 0:
        print(f"\n❌ Error: No data collected. Please check your input files.")
        return
    
    # Create DataFrame
    df_full = pd.DataFrame(all_data)
    
    # Sort by Metric, then Company for easy comparison
    df_full = df_full.sort_values(['Metric', 'Company']).reset_index(drop=True)
    
    # Define column order for full version
    columns_full = [
        'Metric', 'Code', 'Topic', 'Company', 'Industry', 'CID', 'Model', 'Num_Relevant_Pages',
        'Recall@1', 'Recall@5', 'Recall@10', 'Recall@20',
        'nDCG@1', 'nDCG@5', 'nDCG@10', 'nDCG@20',
        'Precision@1', 'Precision@5', 'Precision@10', 'Precision@20',
        'Hit@1', 'Hit@5', 'Hit@10', 'Hit@20'
    ]
    
    # Reorder columns
    df_full = df_full[columns_full]
    
    # Create no-nDCG version by removing nDCG columns
    columns_no_ndcg = [col for col in columns_full if not col.startswith('nDCG')]
    df_no_ndcg = df_full[columns_no_ndcg].copy()
    
    # Print statistics
    print(f"\n📊 Statistics:")
    print(f"   Total rows: {len(df_full)}")
    print(f"   Unique metrics: {df_full['Metric'].nunique()}")
    print(f"   Unique companies: {df_full['Company'].nunique()}")
    print(f"   Industries: {df_full['Industry'].nunique()}")
    
    # Show company list
    print(f"\n   Companies found:")
    for company in sorted(df_full['Company'].unique()):
        count = len(df_full[df_full['Company'] == company])
        print(f"      • {company}: {count} queries")
    
    # Show metrics distribution
    print(f"\n   Metrics by company count:")
    metric_counts = df_full.groupby('Metric')['Company'].nunique().value_counts().sort_index()
    for count, num_metrics in metric_counts.items():
        print(f"      • {num_metrics} metrics appear in {count} company(ies)")
    
    # Save to Excel files
    output_dir.mkdir(parents=True, exist_ok=True)
    
    output_file_full = output_dir / "results_by_metric_full.xlsx"
    output_file_no_ndcg = output_dir / "results_by_metric_no_ndcg.xlsx"
    
    print(f"\n💾 Saving Excel files...")
    
    # Save full version
    with pd.ExcelWriter(output_file_full, engine='openpyxl') as writer:
        df_full.to_excel(writer, sheet_name='Results', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Results']
        for idx, col in enumerate(df_full.columns):
            max_length = max(
                df_full[col].astype(str).map(len).max(),
                len(col)
            )
            # Use proper Excel column letters
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    print(f"   ✓ {output_file_full.name}")
    print(f"      Columns: {len(df_full.columns)}")
    print(f"      Rows: {len(df_full)}")
    
    # Save no-nDCG version
    with pd.ExcelWriter(output_file_no_ndcg, engine='openpyxl') as writer:
        df_no_ndcg.to_excel(writer, sheet_name='Results', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Results']
        for idx, col in enumerate(df_no_ndcg.columns):
            max_length = max(
                df_no_ndcg[col].astype(str).map(len).max(),
                len(col)
            )
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    print(f"   ✓ {output_file_no_ndcg.name}")
    print(f"      Columns: {len(df_no_ndcg.columns)}")
    print(f"      Rows: {len(df_no_ndcg)}")
    
    # Show sample data
    print(f"\n📋 Sample data (first 5 rows):")
    print(df_full[['Metric', 'Company', 'Industry', 'Recall@1', 'Recall@5']].head().to_string())
    
    print(f"\n✅ DONE!")
    print(f"   Full version: {output_file_full}")
    print(f"   No-nDCG version: {output_file_no_ndcg}")
    print(f"{'='*80}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Aggregate evaluation results by metric across companies",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process results_colqwen_eval_mixed
  python aggregate_by_metric.py --input results_colqwen_eval_mixed
  
  # Process with custom output directory
  python aggregate_by_metric.py --input results_colqwen_eval_mixed --output analysis
  
  # Process different result directory
  python aggregate_by_metric.py --input results_retrieval_eval_mixed

Output:
  - results_by_metric_full.xlsx (with nDCG@k metrics)
  - results_by_metric_no_ndcg.xlsx (without nDCG@k metrics)
        """
    )
    
    parser.add_argument(
        '--input',
        default='results_colqwen_eval_mixed',
        help='Input results directory (default: results_colqwen_eval_mixed)'
    )
    
    parser.add_argument(
        '--output',
        default='metric_analysis',
        help='Output directory for Excel files (default: metric_analysis)'
    )
    
    args = parser.parse_args()
    
    # Setup paths
    base_dir = Path(__file__).resolve().parent
    results_dir = base_dir / args.input
    output_dir = base_dir / args.output
    
    # Check if input directory exists
    if not results_dir.exists():
        print(f"\n❌ Error: Input directory not found: {results_dir}")
        print(f"   Please check the directory name and try again.")
        return
    
    # Run aggregation
    aggregate_results_by_metric(results_dir, output_dir)


if __name__ == "__main__":
    main()
